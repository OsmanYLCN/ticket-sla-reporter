"""
Configuration and Rule Definitions for SLA Reporting
"""

import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Target BusinessLine & Severities
TARGET_BUSINESS_LINE = 'Infrastructure Services'
TARGET_SEVERITIES = ['1 - Molto alta', '2 - Alta']

# Regular Expressions for Ticket Classification
NET_KW_PATTERN = re.compile(
    r'\b(network|networking|cato|switch|switches|router|routers|firewall|vpn|wifi|wi-fi|lan|wan|proxy|ise|zpa|zscaler|connessione|reachability|raggiungibilit[aà]|rotte|route|routing|vlan|port|dns)\b',
    re.IGNORECASE
)

HQ_CLOUD_PATTERN = re.compile(
    r'\b(cesano|maderno|milano|milan|tim|hq|headquarters|azure|gcp|cinisello)\b',
    re.IGNORECASE
)

PLANT_PATTERN = re.compile(
    r'\b(kocaeli|izmit|alexandria|alessandria|santo\s*andr[eé]|gravata[ií]|campinas|factory|plant|turkey|egypt|brazil)\b',
    re.IGNORECASE
)

OT_TECH_PATTERN = re.compile(
    r'\b(simplivity|esxi|esx|vm|virtual|server|backup|veeam|proliant|dl380|storage|nas|san|host)\b',
    re.IGNORECASE
)

SAP_PATTERN = re.compile(r'\b(sap|sapsprint)\b', re.IGNORECASE)
SAP_CONN_PATTERN = re.compile(
    r'\b(cato|vpn|network|networking|connessione|connection|reachability|raggiungibilit[aà]|lan|wan|firewall|wifi)\b',
    re.IGNORECASE
)

# Manufacturing Plants & Sites Configuration
PLANT_MAPPINGS = [
    ("Kocaeli Plant (Turkey)", "Kocaeli"),
    ("Alexandria Plant (Egypt)", "Alexandria"),
    ("Santo André Plant (Brazil)", "Santo André"),
    ("Gravataí Plant (Brazil)", "Gravataí"),
    ("Other Network / Remote Sites", "Other Sites")
]

# Corporate Visual Styles & Color Palette
FONT_MAIN_TITLE = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
FONT_SUBTITLE = Font(name='Segoe UI', size=9, italic=True, color='333333')
FONT_SEC_TITLE = Font(name='Segoe UI', size=11, bold=True, color='1F4E78')
FONT_TBL_HEADER = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
FONT_DATA = Font(name='Segoe UI', size=10, color='000000')
FONT_DATA_BOLD = Font(name='Segoe UI', size=10, bold=True, color='000000')
FONT_RESULT = Font(name='Segoe UI', size=11, bold=True, color='1F4E78')

FILL_NAVY = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
FILL_SOFT_HEADER = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
FILL_SUB_BANNER = PatternFill(start_color='EAEEF3', end_color='EAEEF3', fill_type='solid')
FILL_ZEBRA = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
FILL_HIGHLIGHT = PatternFill(start_color='F2F4F8', end_color='F2F4F8', fill_type='solid')

BORDER_THIN = Side(style='thin', color='D9D9D9')
BORDER_DOUBLE = Side(style='double', color='1F4E78')
CELL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
TOTAL_BORDER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_DOUBLE)
