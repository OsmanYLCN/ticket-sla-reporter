import os
import sys
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .config import (
    PLANT_MAPPINGS,
    FONT_MAIN_TITLE,
    FONT_SUBTITLE,
    FONT_SEC_TITLE,
    FONT_TBL_HEADER,
    FONT_DATA,
    FONT_DATA_BOLD,
    FONT_RESULT,
    FILL_NAVY,
    FILL_SOFT_HEADER,
    FILL_SUB_BANNER,
    FILL_ZEBRA,
    FILL_TOTAL,
    FILL_HIGHLIGHT,
    CELL_BORDER,
    TOTAL_BORDER
)
from .classifier import classify_ticket, get_plant_location
from .processor import parse_datetime_value


def check_file_writable(file_path):
    """
    Returns True if file is writable or doesn't exist yet;
    Returns False if locked by an active process (such as Microsoft Excel).
    """
    if not os.path.exists(file_path):
        return True
    try:
        with open(file_path, 'a'):
            pass
        return True
    except (PermissionError, OSError):
        return False


def ensure_file_writable(file_path):
    """
    Prevents PermissionError crashes by verifying write availability.
    If the file is locked in Excel, prompts the user to close it.
    """
    while not check_file_writable(file_path):
        print("\n" + "!" * 75)
        print("  UYARI: HEDEF EXCEL DOSYASI SU ANDA BASKA BIR PROGRAMDA ACIK!")
        print("!" * 75)
        print(f"  Kilitli Dosya : {file_path}")
        print("  Bu dosya su anda Microsoft Excel veya baska bir uygulama tarafindan kullaniliyor.")
        print("  Lutfen acik olan Excel dosyasini KAPATIP devam etmek icin [Enter]'a basin.")
        print("  (Islemi iptal etmek icin 'Q' tuslayip Enter'a basabilirsiniz)")
        try:
            choice = input("\nSeciminiz [Enter: Yeniden Dene, Q: Cikis]: ").strip()
        except KeyboardInterrupt:
            print("\n\n[*] Islem kullanici tarafindan iptal edildi. Cikis yapiliyor.\n")
            sys.exit(0)
            
        if choice.upper() == 'Q':
            print("\n[*] Islem kullanici tarafindan iptal edildi. Cikis yapildi.\n")
            sys.exit(0)


def build_sla_report_workbook(all_records, header_cols, input_path, output_path, total_source_count):
    """
    Creates the output Excel report by copying the full original workbook
    (preserving all quarterly sheets Tickets-Q1..Q4) and populating 'Year-2026'
    and 'Summary' sheets.
    """
    # Verify file is not open/locked before copying
    ensure_file_writable(output_path)
    
    # Copy original file to output path so input file remains 100% untouched
    shutil.copy2(input_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    
    # -------------------------------------------------------------------------
    # 1. Populate 'Year-2026' Consolidated Tab
    # -------------------------------------------------------------------------
    if 'Year-2026' in wb.sheetnames:
        ws_year = wb['Year-2026']
        ws_year.delete_rows(1, ws_year.max_row + 10)
        ws_year.delete_cols(1, ws_year.max_column + 10)
    else:
        ws_year = wb.create_sheet('Year-2026')
        
    ws_year.views.sheetView[0].showGridLines = True
    
    # Headers: 33 original + 6 helper columns
    extended_headers = list(header_cols) + [
        'In_Scope',
        'Resolution_Hours',
        'SLA_Status',
        'Scope_Category',
        'Plant_Site',
        'Quarter'
    ]
    ws_year.append(extended_headers)
    
    # Style Header Row
    for col_idx in range(1, len(extended_headers) + 1):
        cell = ws_year.cell(row=1, column=col_idx)
        cell.font = FONT_TBL_HEADER
        cell.fill = FILL_NAVY
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_year.row_dimensions[1].height = 28
    
    date_col_indices = [3, 10, 11]  # Col C (Creation), Col J (Solution), Col K (Closure)
    
    in_scope_count = 0
    net_count = 0
    ot_count = 0
    passed_count = 0
    failed_count = 0
    
    for idx, rec in enumerate(all_records, start=2):
        in_scope, category = classify_ticket(rec)
        plant_site = get_plant_location(rec)
        
        row_values = []
        for c_idx, h in enumerate(header_cols, start=1):
            val = rec.get(h)
            if c_idx in date_col_indices:
                val = parse_datetime_value(val)
            row_values.append(val)
            
        if in_scope:
            in_scope_count += 1
            if category == 'Network':
                net_count += 1
            elif category == 'Industrial OT':
                ot_count += 1
                
            c_val = row_values[2]  # Col C (Creation)
            s_val = row_values[9]  # Col J (Solution)
            if isinstance(c_val, datetime) and isinstance(s_val, datetime):
                diff_hours = (s_val - c_val).total_seconds() / 3600.0
                if diff_hours <= 4.0:
                    passed_count += 1
                else:
                    failed_count += 1
            
        # Helper Columns:
        # Col 34 (AH): In_Scope
        row_values.append(in_scope)
        # Col 35 (AI): Resolution_Hours: =(J{idx}-C{idx})*24
        row_values.append(f"=(J{idx}-C{idx})*24")
        # Col 36 (AJ): SLA_Status: =IF(AI{idx}<=4, "PASSED", "FAILED")
        row_values.append(f'=IF(AI{idx}<=4, "PASSED", "FAILED")')
        # Col 37 (AK): Scope_Category
        row_values.append(category)
        # Col 38 (AL): Plant_Site
        row_values.append(plant_site)
        # Col 39 (AM): Quarter
        row_values.append(rec.get('_quarter', 'Q1'))
        
        ws_year.append(row_values)
        
        # Date & format adjustments
        for d_col in date_col_indices:
            c_val = ws_year.cell(row=idx, column=d_col).value
            if isinstance(c_val, datetime):
                ws_year.cell(row=idx, column=d_col).number_format = 'yyyy-mm-dd hh:mm:ss'
                
        ws_year.cell(row=idx, column=35).number_format = '0.00'
        ws_year.cell(row=idx, column=34).alignment = Alignment(horizontal='center')
        ws_year.cell(row=idx, column=35).alignment = Alignment(horizontal='right')
        ws_year.cell(row=idx, column=36).alignment = Alignment(horizontal='center')
        
    max_row = len(all_records) + 1
    
    # Auto-adjust column widths for Year-2026
    for col in ws_year.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col[:100])
        ws_year.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws_year.column_dimensions['AH'].width = 14
    ws_year.column_dimensions['AI'].width = 18
    ws_year.column_dimensions['AJ'].width = 14
    ws_year.column_dimensions['AK'].width = 24
    ws_year.column_dimensions['AL'].width = 16
    ws_year.column_dimensions['AM'].width = 12

    # -------------------------------------------------------------------------
    # 2. Build 'Summary' Executive Dashboard Tab
    # -------------------------------------------------------------------------
    if 'Summary' in wb.sheetnames:
        ws_sum = wb['Summary']
        ws_sum.delete_rows(1, ws_sum.max_row + 10)
        ws_sum.delete_cols(1, ws_sum.max_column + 10)
    else:
        ws_sum = wb.create_sheet('Summary', 0)
        
    summary_idx = wb.sheetnames.index('Summary')
    if summary_idx != 0:
        wb._sheets.insert(0, wb._sheets.pop(summary_idx))
        
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Banner (Row 2 & 3) - Spanning B to J for full matrix alignment
    ws_sum.merge_cells('B2:J2')
    cell_title = ws_sum['B2']
    cell_title.value = "PROMETEON TYRE GROUP - IT INFRASTRUCTURE SLA PERFORMANCE REPORT (2026)"
    cell_title.font = FONT_MAIN_TITLE
    cell_title.fill = FILL_NAVY
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 32
    
    ws_sum.merge_cells('B3:J3')
    cell_sub = ws_sum['B3']
    cell_sub.value = f"Scope: BusinessLine = 'Infrastructure Services' | Severity = '1 - Molto alta' & '2 - Alta' | SLA Target: <= 4.0h | Total Analyzed: {total_source_count:,}"
    cell_sub.font = FONT_SUBTITLE
    cell_sub.fill = FILL_SUB_BANNER
    cell_sub.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[3].height = 20

    # -------------------------------------------------------------------------
    # 1. QUARTERLY PERFORMANCE COMPARISON MATRIX (Rows 5 to 12, Columns B to J)
    # -------------------------------------------------------------------------
    ws_sum.merge_cells('B5:J5')
    sec1_banner = ws_sum['B5']
    sec1_banner.value = "1. QUARTERLY SLA PERFORMANCE BREAKDOWN (Q1 - Q4 2026)"
    sec1_banner.font = FONT_SEC_TITLE
    sec1_banner.fill = FILL_SOFT_HEADER
    sec1_banner.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[5].height = 24
    
    # Quarter Group Headers (Row 6)
    ws_sum.merge_cells('B6:B7')
    c_met = ws_sum['B6']
    c_met.value = "Quarterly Metric"
    c_met.font = FONT_TBL_HEADER
    c_met.fill = FILL_NAVY
    c_met.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c_met.border = CELL_BORDER
    ws_sum['B7'].border = CELL_BORDER
    
    quarters = [
        ('Q1', 'Q1 - 2026 (Jan - Mar)', 3, 4),
        ('Q2', 'Q2 - 2026 (Apr - Jun)', 5, 6),
        ('Q3', 'Q3 - 2026 (Jul - Sep)', 7, 8),
        ('Q4', 'Q4 - 2026 (Oct - Dec)', 9, 10),
    ]
    
    for q_code, q_label, c_cnt, c_rate in quarters:
        col_cnt_let = get_column_letter(c_cnt)
        col_rate_let = get_column_letter(c_rate)
        ws_sum.merge_cells(f'{col_cnt_let}6:{col_rate_let}6')
        
        c_q = ws_sum[f'{col_cnt_let}6']
        c_q.value = q_label
        c_q.font = FONT_TBL_HEADER
        c_q.fill = FILL_NAVY
        c_q.alignment = Alignment(horizontal='center', vertical='center')
        c_q.border = CELL_BORDER
        ws_sum[f'{col_rate_let}6'].border = CELL_BORDER
        
        # Subheaders (Row 7)
        c_sub1 = ws_sum.cell(row=7, column=c_cnt, value="Count")
        c_sub1.font = FONT_TBL_HEADER
        c_sub1.fill = FILL_SOFT_HEADER
        c_sub1.alignment = Alignment(horizontal='center', vertical='center')
        c_sub1.border = CELL_BORDER
        
        c_sub2 = ws_sum.cell(row=7, column=c_rate, value="Ratio / Rate")
        c_sub2.font = FONT_TBL_HEADER
        c_sub2.fill = FILL_SOFT_HEADER
        c_sub2.alignment = Alignment(horizontal='center', vertical='center')
        c_sub2.border = CELL_BORDER
        
    ws_sum.row_dimensions[6].height = 22
    ws_sum.row_dimensions[7].height = 20

    # Row 8: Total Critical Tickets (Sev 1 + Sev 2)
    ws_sum['B8'] = "Total Critical Tickets (Sev 1 + Sev 2)"
    for q_code, _, c_cnt, c_rate in quarters:
        ws_sum.cell(row=8, column=c_cnt, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AM$2:$AM${max_row}, "{q_code}")')
        col_cnt_let = get_column_letter(c_cnt)
        ws_sum.cell(row=8, column=c_rate, value=f'=IF({col_cnt_let}8>0, 1.0, "-")')

    # Row 9: Closed Within 4 Hours (SLA Passed)
    ws_sum['B9'] = "Closed Within 4 Hours (SLA Passed)"
    for q_code, _, c_cnt, c_rate in quarters:
        ws_sum.cell(row=9, column=c_cnt, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AM$2:$AM${max_row}, "{q_code}", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")')
        col_cnt_let = get_column_letter(c_cnt)
        ws_sum.cell(row=9, column=c_rate, value=f'=IF({col_cnt_let}8>0, {col_cnt_let}9/{col_cnt_let}8, "-")')

    # Row 10: Exceeding 4 Hours / SLA Breach (Failed)
    ws_sum['B10'] = "Exceeding 4 Hours / SLA Breach (Failed)"
    for q_code, _, c_cnt, c_rate in quarters:
        ws_sum.cell(row=10, column=c_cnt, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AM$2:$AM${max_row}, "{q_code}", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")')
        col_cnt_let = get_column_letter(c_cnt)
        ws_sum.cell(row=10, column=c_rate, value=f'=IF({col_cnt_let}8>0, {col_cnt_let}10/{col_cnt_let}8, "-")')

    # Row 11: KPI Target
    ws_sum['B11'] = "KPI Target"
    for _, _, c_cnt, c_rate in quarters:
        ws_sum.cell(row=11, column=c_cnt, value="-")
        ws_sum.cell(row=11, column=c_rate, value=0.90)

    # Row 12: Result
    ws_sum['B12'] = "Quarterly Result"
    for _, _, c_cnt, c_rate in quarters:
        col_cnt_let = get_column_letter(c_cnt)
        col_rate_let = get_column_letter(c_rate)
        ws_sum.merge_cells(f'{col_cnt_let}12:{col_rate_let}12')
        res_cell = ws_sum[f'{col_cnt_let}12']
        res_cell.value = f'=IF({col_cnt_let}8=0, "NO DATA", IF({col_rate_let}9>={col_rate_let}11, "PASSED", "FAILED"))'
        ws_sum[f'{col_rate_let}12'].border = CELL_BORDER

    # Styling Table 1 (Quarterly Matrix)
    for r in range(8, 13):
        ws_sum.row_dimensions[r].height = 22
        is_result = (r == 12)
        is_total = (r == 8)
        
        for col_idx in range(2, 11):
            cell = ws_sum.cell(row=r, column=col_idx)
            cell.border = TOTAL_BORDER if is_total else CELL_BORDER
            
            if is_result:
                cell.fill = FILL_HIGHLIGHT
                cell.font = FONT_RESULT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = FONT_DATA_BOLD if is_total else FONT_DATA
                if r % 2 == 1:
                    cell.fill = FILL_ZEBRA
                
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                elif col_idx in [3, 5, 7, 9]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        cell.number_format = '#,##0'
                elif col_idx in [4, 6, 8, 10]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if r in [8, 9, 10, 11] and cell.value != "-":
                        cell.number_format = '0.0%'

    # -------------------------------------------------------------------------
    # 2. FULL YEAR 2026 CONSOLIDATED PERFORMANCE (Rows 15 to 21)
    # -------------------------------------------------------------------------
    ws_sum.merge_cells('B14:E14')
    sec2_banner = ws_sum['B14']
    sec2_banner.value = "2. FULL YEAR 2026 - CONSOLIDATED INFRASTRUCTURE SLA PERFORMANCE"
    sec2_banner.font = FONT_SEC_TITLE
    sec2_banner.fill = FILL_SOFT_HEADER
    sec2_banner.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[14].height = 24
    
    headers_t2 = ['Metric', 'Count / Value', 'Ratio / Rate', 'Status']
    for col_idx, h_text in enumerate(headers_t2, start=2):
        c = ws_sum.cell(row=15, column=col_idx)
        c.value = h_text
        c.font = FONT_TBL_HEADER
        c.fill = FILL_NAVY
        c.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
        c.border = CELL_BORDER
    ws_sum.row_dimensions[15].height = 24
    
    # Row 16: Total Critical Tickets
    ws_sum['B16'] = "Total Critical Tickets (Sev 1 + Sev 2)"
    ws_sum['C16'] = f"=COUNTIF('Year-2026'!$AH$2:$AH${max_row}, TRUE)"
    ws_sum['D16'] = 1.0
    ws_sum['E16'] = "-"
    
    # Row 17: Closed Within 4 Hours (SLA Passed)
    ws_sum['B17'] = "Closed Within 4 Hours (SLA Passed)"
    ws_sum['C17'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")'
    ws_sum['D17'] = "=IF(C16>0, C17/C16, 0)"
    ws_sum['E17'] = "-"
    
    # Row 18: Exceeding 4 Hours / SLA Breach (Failed)
    ws_sum['B18'] = "Exceeding 4 Hours / SLA Breach (Failed)"
    ws_sum['C18'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")'
    ws_sum['D18'] = "=IF(C16>0, C18/C16, 0)"
    ws_sum['E18'] = "-"
    
    # Row 19: KPI Target
    ws_sum['B19'] = "KPI Target"
    ws_sum['C19'] = "-"
    ws_sum['D19'] = 0.90
    ws_sum['E19'] = "-"
    
    # Row 20: Result
    ws_sum['B20'] = "Result"
    ws_sum['C20'] = "-"
    ws_sum['D20'] = "-"
    ws_sum['E20'] = '=IF(D17>=D19, "PASSED", "FAILED")'
    
    # Styling Table 2 (Consolidated Annual Table)
    for r in range(16, 21):
        ws_sum.row_dimensions[r].height = 22
        for col_idx in range(2, 6):
            cell = ws_sum.cell(row=r, column=col_idx)
            cell.font = FONT_DATA_BOLD if r in [16, 20] else FONT_DATA
            cell.border = CELL_BORDER
            if r == 20:
                cell.fill = FILL_HIGHLIGHT
                if col_idx == 5:
                    cell.font = FONT_RESULT
            elif r % 2 == 1:
                cell.fill = FILL_ZEBRA
                
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.number_format = '#,##0'
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if r in [16, 17, 18, 19] and cell.value != "-":
                    cell.number_format = '0.0%'
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # -------------------------------------------------------------------------
    # 3. DOMAIN BREAKDOWN (Rows 23 to 27)
    # -------------------------------------------------------------------------
    ws_sum.merge_cells('B22:G22')
    sec3_banner = ws_sum['B22']
    sec3_banner.value = "3. DOMAIN & INFRASTRUCTURE CATEGORY BREAKDOWN"
    sec3_banner.font = FONT_SEC_TITLE
    sec3_banner.fill = FILL_SOFT_HEADER
    sec3_banner.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[22].height = 24
    
    headers_t3 = ['Category', 'In-Scope Incidents', 'SLA Met (Passed)', 'SLA Breached (Failed)', 'SLA Compliance (%)', 'Avg Resolution (Hours)']
    for col_idx, h_text in enumerate(headers_t3, start=2):
        c = ws_sum.cell(row=23, column=col_idx)
        c.value = h_text
        c.font = FONT_TBL_HEADER
        c.fill = FILL_NAVY
        c.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
        c.border = CELL_BORDER
    ws_sum.row_dimensions[23].height = 24
    
    # Row 24: Network
    ws_sum['B24'] = "Managed Network Services"
    ws_sum['C24'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Network")'
    ws_sum['D24'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Network", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")'
    ws_sum['E24'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Network", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")'
    ws_sum['F24'] = "=IF(C24>0, D24/C24, 0)"
    ws_sum['G24'] = f'=AVERAGEIFS(\'Year-2026\'!$AI$2:$AI${max_row}, \'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Network")'
    
    # Row 25: Industrial OT
    ws_sum['B25'] = "Industrial OT & Plant Infrastructure"
    ws_sum['C25'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Industrial OT")'
    ws_sum['D25'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Industrial OT", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")'
    ws_sum['E25'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Industrial OT", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")'
    ws_sum['F25'] = "=IF(C25>0, D25/C25, 0)"
    ws_sum['G25'] = f'=AVERAGEIFS(\'Year-2026\'!$AI$2:$AI${max_row}, \'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AK$2:$AK${max_row}, "Industrial OT")'
    
    # Row 26: Total
    ws_sum['B26'] = "Total In-Scope Infrastructure"
    ws_sum['C26'] = "=SUM(C24:C25)"
    ws_sum['D26'] = "=SUM(D24:D25)"
    ws_sum['E26'] = "=SUM(E24:E25)"
    ws_sum['F26'] = "=IF(C26>0, D26/C26, 0)"
    ws_sum['G26'] = f'=AVERAGEIFS(\'Year-2026\'!$AI$2:$AI${max_row}, \'Year-2026\'!$AH$2:$AH${max_row}, TRUE)'
    
    for r in range(24, 27):
        ws_sum.row_dimensions[r].height = 20
        is_tot = (r == 26)
        for col_idx in range(2, 8):
            cell = ws_sum.cell(row=r, column=col_idx)
            cell.font = FONT_DATA_BOLD if is_tot else FONT_DATA
            cell.border = TOTAL_BORDER if is_tot else CELL_BORDER
            if is_tot:
                cell.fill = FILL_TOTAL
            elif r % 2 == 1:
                cell.fill = FILL_ZEBRA
                
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            elif col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.0%'
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.00'

    # -------------------------------------------------------------------------
    # 4. SEVERITY BREAKDOWN (Rows 29 to 33)
    # -------------------------------------------------------------------------
    ws_sum.merge_cells('B28:F28')
    sec4_banner = ws_sum['B28']
    sec4_banner.value = "4. BREAKDOWN BY SEVERITY LEVEL"
    sec4_banner.font = FONT_SEC_TITLE
    sec4_banner.fill = FILL_SOFT_HEADER
    sec4_banner.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[28].height = 24
    
    headers_t4 = ['Severity Level', 'In-Scope Incidents', 'SLA Met (Passed)', 'SLA Breached (Failed)', 'SLA Compliance (%)']
    for col_idx, h_text in enumerate(headers_t4, start=2):
        c = ws_sum.cell(row=29, column=col_idx)
        c.value = h_text
        c.font = FONT_TBL_HEADER
        c.fill = FILL_NAVY
        c.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
        c.border = CELL_BORDER
    ws_sum.row_dimensions[29].height = 24
    
    # Row 30: Severity 1
    ws_sum['B30'] = "Severity 1 - Molto alta (Critical)"
    ws_sum['C30'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "1 - Molto alta")'
    ws_sum['D30'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "1 - Molto alta", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")'
    ws_sum['E30'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "1 - Molto alta", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")'
    ws_sum['F30'] = "=IF(C30>0, D30/C30, 0)"
    
    # Row 31: Severity 2
    ws_sum['B31'] = "Severity 2 - Alta (High)"
    ws_sum['C31'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "2 - Alta")'
    ws_sum['D31'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "2 - Alta", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")'
    ws_sum['E31'] = f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$I$2:$I${max_row}, "2 - Alta", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")'
    ws_sum['F31'] = "=IF(C31>0, D31/C31, 0)"
    
    # Row 32: Total
    ws_sum['B32'] = "Total Severity 1 & 2"
    ws_sum['C32'] = "=SUM(C30:C31)"
    ws_sum['D32'] = "=SUM(D30:D31)"
    ws_sum['E32'] = "=SUM(E30:E31)"
    ws_sum['F32'] = "=IF(C32>0, D32/C32, 0)"
    
    for r in range(30, 33):
        ws_sum.row_dimensions[r].height = 20
        is_tot = (r == 32)
        for col_idx in range(2, 7):
            cell = ws_sum.cell(row=r, column=col_idx)
            cell.font = FONT_DATA_BOLD if is_tot else FONT_DATA
            cell.border = TOTAL_BORDER if is_tot else CELL_BORDER
            if is_tot:
                cell.fill = FILL_TOTAL
            elif r % 2 == 1:
                cell.fill = FILL_ZEBRA
                
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            elif col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.0%'

    # -------------------------------------------------------------------------
    # 5. PLANT BREAKDOWN (Rows 35 to 42)
    # -------------------------------------------------------------------------
    ws_sum.merge_cells('B34:F34')
    sec5_banner = ws_sum['B34']
    sec5_banner.value = "5. BREAKDOWN BY MANUFACTURING PLANT & LOCATION"
    sec5_banner.font = FONT_SEC_TITLE
    sec5_banner.fill = FILL_SOFT_HEADER
    sec5_banner.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_sum.row_dimensions[34].height = 24
    
    headers_t5 = ['Manufacturing Plant / Site', 'In-Scope Incidents', 'SLA Met (Passed)', 'SLA Breached (Failed)', 'SLA Compliance (%)']
    for col_idx, h_text in enumerate(headers_t5, start=2):
        c = ws_sum.cell(row=35, column=col_idx)
        c.value = h_text
        c.font = FONT_TBL_HEADER
        c.fill = FILL_NAVY
        c.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
        c.border = CELL_BORDER
    ws_sum.row_dimensions[35].height = 24
    
    start_r = 36
    for idx, (p_label, p_code) in enumerate(PLANT_MAPPINGS):
        curr_r = start_r + idx
        ws_sum.cell(row=curr_r, column=2, value=p_label)
        ws_sum.cell(row=curr_r, column=3, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AL$2:$AL${max_row}, "{p_code}")')
        ws_sum.cell(row=curr_r, column=4, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AL$2:$AL${max_row}, "{p_code}", \'Year-2026\'!$AJ$2:$AJ${max_row}, "PASSED")')
        ws_sum.cell(row=curr_r, column=5, value=f'=COUNTIFS(\'Year-2026\'!$AH$2:$AH${max_row}, TRUE, \'Year-2026\'!$AL$2:$AL${max_row}, "{p_code}", \'Year-2026\'!$AJ$2:$AJ${max_row}, "FAILED")')
        ws_sum.cell(row=curr_r, column=6, value=f'=IF(C{curr_r}>0, D{curr_r}/C{curr_r}, 0)')
        
    tot_plant_r = start_r + len(PLANT_MAPPINGS)
    ws_sum.cell(row=tot_plant_r, column=2, value="Total Plants & Sites")
    ws_sum.cell(row=tot_plant_r, column=3, value=f"=SUM(C{start_r}:C{tot_plant_r-1})")
    ws_sum.cell(row=tot_plant_r, column=4, value=f"=SUM(D{start_r}:D{tot_plant_r-1})")
    ws_sum.cell(row=tot_plant_r, column=5, value=f"=SUM(E{start_r}:E{tot_plant_r-1})")
    ws_sum.cell(row=tot_plant_r, column=6, value=f"=IF(C{tot_plant_r}>0, D{tot_plant_r}/C{tot_plant_r}, 0)")
    
    for r in range(start_r, tot_plant_r + 1):
        ws_sum.row_dimensions[r].height = 20
        is_tot = (r == tot_plant_r)
        for col_idx in range(2, 7):
            cell = ws_sum.cell(row=r, column=col_idx)
            cell.font = FONT_DATA_BOLD if is_tot else FONT_DATA
            cell.border = TOTAL_BORDER if is_tot else CELL_BORDER
            if is_tot:
                cell.fill = FILL_TOTAL
            elif r % 2 == 1:
                cell.fill = FILL_ZEBRA
                
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            elif col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.0%'

    # Set Column Widths for Summary (Columns A through J)
    ws_sum.column_dimensions['A'].width = 4
    ws_sum.column_dimensions['B'].width = 44
    ws_sum.column_dimensions['C'].width = 16
    ws_sum.column_dimensions['D'].width = 16
    ws_sum.column_dimensions['E'].width = 16
    ws_sum.column_dimensions['F'].width = 16
    ws_sum.column_dimensions['G'].width = 16
    ws_sum.column_dimensions['H'].width = 16
    ws_sum.column_dimensions['I'].width = 16
    ws_sum.column_dimensions['J'].width = 16
    while True:
        try:
            ensure_file_writable(output_path)
            wb.save(output_path)
            wb.close()
            break
        except (PermissionError, OSError):
            ensure_file_writable(output_path)
            
    stats = {
        'in_scope_count': in_scope_count,
        'max_row': max_row,
        'net_count': net_count,
        'ot_count': ot_count,
        'passed_count': passed_count,
        'failed_count': failed_count
    }
    return stats
