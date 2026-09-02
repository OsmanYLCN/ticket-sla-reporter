"""
Data ingestion, date normalization, and quarterly ticket consolidation.
"""

from datetime import datetime
import openpyxl


def parse_datetime_value(val):
    """
    Parses a date string or returns datetime object for Excel serialization.
    """
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    val_str = str(val).strip()
    for fmt in (
        '%d-%m-%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%d-%m-%Y %H:%M',
        '%d/%m/%Y %H:%M'
    ):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return val_str


def load_and_consolidate_tickets(file_path):
    """
    Reads all quarterly ticket sheets (Q1..Q4) from source workbook in read-only mode.
    Returns: (all_records, header_cols, quarterly_counts)
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    quarter_sheets = ['Tickets-Q1', 'Tickets-Q2', 'Tickets-Q3', 'Tickets-Q4']
    
    header_cols = None
    all_records = []
    quarterly_counts = {}
    
    for q_name in quarter_sheets:
        if q_name not in wb.sheetnames:
            continue
            
        ws_q = wb[q_name]
        
        # Read header and rows
        rows_iter = ws_q.iter_rows(values_only=True)
        try:
            first_row = next(rows_iter)
        except StopIteration:
            continue
            
        if not first_row or not any(first_row):
            continue
            
        if header_cols is None:
            header_cols = list(first_row)
            
        col_names = [str(c) if c is not None else f"Col_{i+1}" for i, c in enumerate(header_cols)]
        
        q_count = 0
        for row in rows_iter:
            if not any(row):
                continue
            row_dict = {col_names[i]: row[i] if i < len(row) else None for i in range(len(col_names))}
            row_dict['_quarter'] = q_name.replace('Tickets-', '')
            all_records.append(row_dict)
            q_count += 1
            
        quarterly_counts[q_name] = q_count
        
    wb.close()
    return all_records, header_cols, quarterly_counts
