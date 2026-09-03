"""
Unit Tests for src.processor
============================
Tests datetime parsing, multi-quarter workbook ingestion, and data consolidation.
"""

import unittest
import os
import tempfile
from datetime import datetime
import openpyxl

from src.processor import parse_datetime_value, load_and_consolidate_tickets


class TestProcessor(unittest.TestCase):

    def test_parse_datetime_none_and_empty(self):
        """None and empty strings must return None."""
        self.assertIsNone(parse_datetime_value(None))
        self.assertIsNone(parse_datetime_value(""))

    def test_parse_datetime_passthrough(self):
        """Existing datetime instances should be returned as-is."""
        now = datetime.now()
        self.assertEqual(parse_datetime_value(now), now)

    def test_parse_datetime_formats(self):
        """Tests parsing across various standard enterprise timestamp formats."""
        cases = [
            ("31-08-2026 14:30:15", datetime(2026, 8, 31, 14, 30, 15)),
            ("2026-08-31 14:30:15", datetime(2026, 8, 31, 14, 30, 15)),
            ("31/08/2026 14:30:15", datetime(2026, 8, 31, 14, 30, 15)),
            ("2026/08/31 14:30:15", datetime(2026, 8, 31, 14, 30, 15)),
            ("31-08-2026 14:30", datetime(2026, 8, 31, 14, 30)),
            ("31/08/2026 14:30", datetime(2026, 8, 31, 14, 30))
        ]
        for val_str, expected in cases:
            with self.subTest(val_str=val_str):
                self.assertEqual(parse_datetime_value(val_str), expected)

    def test_parse_datetime_invalid_fallback(self):
        """Unparseable string values should be safely returned as string without crashing."""
        self.assertEqual(parse_datetime_value("NOT_A_DATE"), "NOT_A_DATE")

    def test_load_and_consolidate_tickets(self):
        """Tests reading and consolidating multiple quarter sheets from an Excel workbook."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # 1. Create Tickets-Q1 with 2 records
        ws_q1 = wb.create_sheet('Tickets-Q1')
        ws_q1.append(['Ticket_ID', 'BusinessLine', 'Severity', 'Oggetto'])
        ws_q1.append(['INC001', 'Infrastructure Services', '1 - Molto alta', 'Switch down'])
        ws_q1.append(['INC002', 'Infrastructure Services', '2 - Alta', 'VPN down'])

        # 2. Create Tickets-Q2 with 1 record
        ws_q2 = wb.create_sheet('Tickets-Q2')
        ws_q2.append(['Ticket_ID', 'BusinessLine', 'Severity', 'Oggetto'])
        ws_q2.append(['INC003', 'Infrastructure Services', '1 - Molto alta', 'Router alert'])

        # 3. Leave Tickets-Q3 absent (to test missing sheet handling)
        # 4. Create empty Tickets-Q4 (header only, to test empty sheet handling)
        ws_q4 = wb.create_sheet('Tickets-Q4')
        ws_q4.append(['Ticket_ID', 'BusinessLine', 'Severity', 'Oggetto'])

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            wb.save(tmp_path)
            wb.close()

            records, headers, counts = load_and_consolidate_tickets(tmp_path)

            self.assertEqual(len(records), 3)
            self.assertEqual(headers, ['Ticket_ID', 'BusinessLine', 'Severity', 'Oggetto'])
            self.assertEqual(counts['Tickets-Q1'], 2)
            self.assertEqual(counts['Tickets-Q2'], 1)
            self.assertEqual(counts['Tickets-Q3'], 0)
            self.assertEqual(counts['Tickets-Q4'], 0)

            # Validate quarter attribution
            self.assertEqual(records[0]['_quarter'], 'Q1')
            self.assertEqual(records[1]['_quarter'], 'Q1')
            self.assertEqual(records[2]['_quarter'], 'Q2')

        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)


if __name__ == '__main__':
    unittest.main()
