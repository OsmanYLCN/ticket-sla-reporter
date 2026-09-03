"""
Unit Tests for src.excel_builder
================================
Tests file locking detection, Excel report construction, dynamic formula generation,
and executive Summary dashboard formatting.
"""

import unittest
import os
import tempfile
from datetime import datetime, timedelta
import openpyxl

from src.excel_builder import check_file_writable, build_sla_report_workbook


class TestExcelBuilder(unittest.TestCase):

    def test_check_file_writable_non_existent(self):
        """A file path that does not exist yet should be considered writable."""
        fake_path = os.path.join(tempfile.gettempdir(), "non_existent_sla_test_file.xlsx")
        if os.path.exists(fake_path):
            os.remove(fake_path)
        self.assertTrue(check_file_writable(fake_path))

    def test_check_file_writable_existing(self):
        """A standard closed file should be considered writable."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            self.assertTrue(check_file_writable(tmp_path))
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_build_sla_report_workbook(self):
        """Tests end-to-end report workbook generation, sheet ordering, and formula bindings."""
        # Create minimal source workbook
        wb_src = openpyxl.Workbook()
        ws_q1 = wb_src.active
        ws_q1.title = 'Tickets-Q1'
        
        # 33 standard headers
        headers = [
            'ID', 'Codice', 'Data Creazione', 'Stato', 'Chiusura', 'Riapertura',
            'BusinessLine', 'Descrizione BusinessLine', 'Severity', 'Data Soluzione',
            'Data Risoluzione', 'Servizio', 'Tipologia Calcolata', 'Oggetto',
            'Configuration Item', 'Descrizione Articolo', 'LocalitaCI',
            'Ubicazione sede Utente', 'Assegnatario', 'Gruppo', 'Solutore',
            'Priorita', 'Impatto', 'Urgenza', 'Categoria', 'Sottocategoria',
            'SLA Scadenza', 'SLA Target', 'Note', 'Risoluzione', 'Note Chiusura',
            'Data Ultima Modifica', 'Utente Ultima Modifica'
        ]
        ws_q1.append(headers)

        now = datetime(2026, 3, 15, 10, 0, 0)
        # Record 1: In-Scope Network, Passed (resolved in 2 hours)
        rec1 = {h: f"val_{i}" for i, h in enumerate(headers)}
        rec1['BusinessLine'] = 'Infrastructure Services'
        rec1['Severity'] = '1 - Molto alta'
        rec1['Servizio'] = 'MANAGED NETWORK SERVICE'
        rec1['Oggetto'] = 'Core switch down'
        rec1['Data Creazione'] = now
        rec1['Data Soluzione'] = now + timedelta(hours=2)
        rec1['LocalitaCI'] = 'Kocaeli'
        rec1['Ubicazione sede Utente'] = 'Kocaeli'
        rec1['_quarter'] = 'Q1'

        # Record 2: In-Scope Network, Failed (resolved in 6 hours)
        rec2 = {h: f"val_{i}" for i, h in enumerate(headers)}
        rec2['BusinessLine'] = 'Infrastructure Services'
        rec2['Severity'] = '2 - Alta'
        rec2['Servizio'] = 'MANAGED NETWORK SERVICE'
        rec2['Oggetto'] = 'VPN gateway error'
        rec2['Data Creazione'] = now
        rec2['Data Soluzione'] = now + timedelta(hours=6)
        rec2['LocalitaCI'] = 'Alexandria'
        rec2['Ubicazione sede Utente'] = 'Alexandria'
        rec2['_quarter'] = 'Q1'

        # Record 3: Out of Scope (BusinessLine = Application)
        rec3 = {h: f"val_{i}" for i, h in enumerate(headers)}
        rec3['BusinessLine'] = 'Application Services'
        rec3['Severity'] = '1 - Molto alta'
        rec3['Data Creazione'] = now
        rec3['Data Soluzione'] = now + timedelta(hours=1)
        rec3['_quarter'] = 'Q1'

        all_records = [rec1, rec2, rec3]
        for r in all_records:
            ws_q1.append([r.get(h) for h in headers])

        # Add empty Year-2026 sheet
        wb_src.create_sheet('Year-2026')

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_in:
            in_path = tmp_in.name
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_out:
            out_path = tmp_out.name

        try:
            wb_src.save(in_path)
            wb_src.close()

            # Execute workbook builder
            stats = build_sla_report_workbook(all_records, headers, in_path, out_path, total_source_count=3)

            # Validate returned metrics
            self.assertEqual(stats['in_scope_count'], 2)
            self.assertEqual(stats['net_count'], 2)
            self.assertEqual(stats['ot_count'], 0)
            self.assertEqual(stats['passed_count'], 1)
            self.assertEqual(stats['failed_count'], 1)
            self.assertEqual(stats['max_row'], 4)  # header + 3 records

            # Inspect generated workbook
            wb_res = openpyxl.load_workbook(out_path, data_only=False)
            
            # Sheet Ordering: Summary must be first
            self.assertEqual(wb_res.sheetnames[0], 'Summary')
            self.assertIn('Year-2026', wb_res.sheetnames)

            ws_sum = wb_res['Summary']
            # Table 1: Quarterly matrix formula in C8 (Total Critical Q1)
            self.assertIn("COUNTIFS('Year-2026'!$AH$2:$AH$4, TRUE, 'Year-2026'!$AM$2:$AM$4, \"Q1\")", ws_sum['C8'].value)
            
            # Table 2: Full Year table formula in C16 (Total Critical Year)
            self.assertIn("COUNTIF('Year-2026'!$AH$2:$AH$4, TRUE)", ws_sum['C16'].value)

            # Table 2: Closed Within 4h in C17
            self.assertIn("COUNTIFS('Year-2026'!$AH$2:$AH$4, TRUE, 'Year-2026'!$AJ$2:$AJ$4, \"PASSED\")", ws_sum['C17'].value)

            # Year-2026 sheet validation
            ws_yr = wb_res['Year-2026']
            self.assertEqual(ws_yr.max_row, 4)
            # Check helper columns in Row 2
            self.assertTrue(ws_yr.cell(row=2, column=34).value)  # In_Scope
            self.assertEqual(ws_yr.cell(row=2, column=37).value, 'Network')  # Scope_Category
            self.assertEqual(ws_yr.cell(row=2, column=38).value, 'Kocaeli')  # Plant_Site
            self.assertEqual(ws_yr.cell(row=2, column=39).value, 'Q1')  # Quarter

            wb_res.close()

        finally:
            if os.path.exists(in_path):
                os.remove(in_path)
            if os.path.exists(out_path):
                os.remove(out_path)


if __name__ == '__main__':
    unittest.main()
