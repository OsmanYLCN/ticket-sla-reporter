"""
Integration Tests for Ticket SLA Reporter
=========================================
Performs end-to-end processing against real company sample data and CLI UI components.
"""

import unittest
import os
import tempfile
import openpyxl

from src.processor import load_and_consolidate_tickets
from src.excel_builder import build_sla_report_workbook
from generate_sla_summary import UI


class TestIntegration(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.sample_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'sample_data',
            'tickets_export_20260831_084937.xlsx'
        )

    def test_ui_vlen_and_pad(self):
        """Tests that UI padding properly ignores non-printable ANSI escape sequences."""
        raw_text = "Test Message"
        ansi_text = f"{UI.GREEN}{raw_text}{UI.RESET}"
        
        self.assertEqual(UI.vlen(raw_text), 12)
        self.assertEqual(UI.vlen(ansi_text), 12)

        padded = UI.pad(ansi_text, 20, align='left')
        self.assertEqual(UI.vlen(padded), 20)
        self.assertTrue(padded.endswith(' ' * 8))

        padded_center = UI.pad(ansi_text, 20, align='center')
        self.assertEqual(UI.vlen(padded_center), 20)

    def test_full_pipeline_on_sample_data(self):
        """Runs the entire pipeline against sample_data and verifies audit accuracy."""
        if not os.path.exists(self.sample_path):
            self.skipTest(f"Sample data file not found at {self.sample_path}")

        # Step 1: Ingestion
        all_records, headers, quarterly_counts = load_and_consolidate_tickets(self.sample_path)
        
        self.assertEqual(len(all_records), 10002)
        self.assertEqual(quarterly_counts['Tickets-Q1'], 5398)
        self.assertEqual(quarterly_counts['Tickets-Q2'], 4604)
        self.assertEqual(quarterly_counts['Tickets-Q3'], 0)
        self.assertEqual(quarterly_counts['Tickets-Q4'], 0)

        # Step 2: Generation into temporary output
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_output = tmp.name

        try:
            stats = build_sla_report_workbook(
                all_records,
                headers,
                self.sample_path,
                tmp_output,
                total_source_count=len(all_records)
            )

            # Mathematical Audit Assertions
            self.assertEqual(stats['in_scope_count'], 85)
            self.assertEqual(stats['net_count'], 63)
            self.assertEqual(stats['ot_count'], 22)
            self.assertEqual(stats['passed_count'], 32)
            self.assertEqual(stats['failed_count'], 53)
            self.assertEqual(stats['max_row'], 10003)

            # Compliance Rate calculation: 32 / 85 = 37.647...%
            compliance_rate = stats['passed_count'] / stats['in_scope_count'] * 100
            self.assertAlmostEqual(compliance_rate, 37.647, places=2)

            # Verify workbook can be read cleanly
            wb = openpyxl.load_workbook(tmp_output, read_only=True)
            self.assertIn('Summary', wb.sheetnames)
            self.assertIn('Year-2026', wb.sheetnames)
            self.assertIn('Tickets-Q1', wb.sheetnames)
            self.assertIn('Tickets-Q2', wb.sheetnames)
            self.assertEqual(wb.sheetnames[0], 'Summary')
            wb.close()

        finally:
            if os.path.exists(tmp_output):
                os.remove(tmp_output)


if __name__ == '__main__':
    unittest.main()
